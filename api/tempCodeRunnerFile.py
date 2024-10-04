from paddleocr import PaddleOCR
import os
import openpyxl #thư viện đọc excel
from unidecode import unidecode  #thư viện để in hoa bỏ dấu

ocr = PaddleOCR(use_angle_cls=True, lang='vi')  # Đặt ngôn ngữ là tiếng Việt

#####thay thế hàm đọc ảnh bằng đọc từ excel
#img_path = r"D:\Edu\Python\StudentID_FaceVerification\student-id-face-matching\List of candidates\0.png"
excel_path = r'D:\Edu\Python\StudentID_FaceVerification\student-id-face-matching\api\List of candidates\DemoDanhSach.xlsx'

###### Sẽ thay thế bằng hàm đọc Excel
#hàm đọc ảnh
# def read_from_image(img_path):
#     if os.path.exists(img_path):
#         # Nhận diện văn bản từ file ảnh
#         result = ocr.ocr(img_path, cls=True)

#         # Trích xuất văn bản từ kết quả OCR
#         extracted_names = []
#         names_and_ids = []

#         for line in result[0]:  # Xử lý từng dòng được OCR nhận diện
#             text = line[1][0].strip()
#             names_and_ids.append(text)

#         # Xử lý để lấy tên và mã sinh viên
#         for i in range(1, len(names_and_ids)):  # Bắt đầu từ dòng thứ 2
#             current_line = names_and_ids[i]
#             if len(current_line) == 15 and current_line.isdigit():  # Nếu dòng hiện tại là mã sinh viên
#                 student_id = current_line
#                 student_name = names_and_ids[i - 1]  # Lấy tên từ dòng trên
#                 # Loại bỏ dấu và chuyển thành in hoa
#                 student_name = unidecode(student_name).upper()
#                 extracted_names.append(f"{student_name} - {student_id}")

#         return extracted_names
#     else:
#         print(f"File {img_path} không tồn tại.")
#         return []

# Hàm phát hiện các cột chứa "Họ tên" và "Mã sinh viên"
def find_columns(sheet):
    name_col = None
    msv_col = None
    header_row = None

    # Duyệt qua các dòng để tìm dòng tiêu đề
    for row in sheet.iter_rows(values_only=True):
        for idx, cell_value in enumerate(row):
            if cell_value:
                # Tìm cột có chứa "Họ tên"
                if "họ tên" in str(cell_value).lower():
                    name_col = idx + 1  # openpyxl đánh số cột từ 1
                # Tìm cột có chứa "Mã sinh viên"
                elif "mã sinh viên" in str(cell_value).lower():
                    msv_col = idx + 1

        # Nếu đã tìm thấy cả 2 cột thì dừng lại
        if name_col and msv_col:
            header_row = row
            break

    return name_col, msv_col, header_row

#hàm đọc danh sách sinh viên từ Excel
def read_from_excel(excel_path):
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # Tìm vị trí cột chứa "Họ tên" và "Mã sinh viên"
        name_col, msv_col, header_row = find_columns(sheet)

        if name_col and msv_col:
            excel_data = []
            # Bắt đầu đọc từ dòng sau tiêu đề
            for row in sheet.iter_rows(min_row=sheet.min_row + 1, values_only=True):
                name = row[name_col - 1]  # Trừ 1 vì openpyxl đánh số từ 1
                msv = row[msv_col - 1]
                if name and msv:  # Chỉ lấy những dòng có dữ liệu
                    # Loại bỏ dấu và chuyển thành in hoa
                    name = unidecode(name).upper()
                    excel_data.append(f"{name} - {msv}")
            return excel_data
        else:
            print("Không tìm thấy cột 'Họ tên' và 'Mã sinh viên'.")
            return []
    else:
        print(f"File Excel {excel_path} không tồn tại.")
        return []

##### bỏ phần lấy từ ảnh
# Đọc danh sách từ ảnh và Excel
#students_from_image = read_from_image(img_path)
students_from_excel = read_from_excel(excel_path)

########Thay thế bằng đọc qua Excel và lưu vào DB
#hiện kết quả đọc ảnh
# if students_from_image:
#     print("Danh sách sinh viên từ ảnh:")
#     for student in students_from_image:
#         print(student)

#     #lưu vào file txt
#     with open(r'D:\Edu\Python\StudentID_FaceVerification\student-id-face-matching\List of candidates\extracted_list\student_list_from_image.txt', 'w', encoding='utf-8') as f:
#         for student in students_from_image:
#             f.write(student + '\n')
#     print("Danh sách sinh viên từ ảnh đã được lưu vào file 'student_list_from_image.txt'.")

#hiển thị kết quả và lưu danh sách sinh viên từ Excel
if students_from_excel:
    print("Danh sách sinh viên từ Excel:")
    for student in students_from_excel:
        print(student)

    ##### Sẽ thay thế bằng cách lưu vào DB
    #lưu sang txt
    with open(r'D:\Edu\Python\StudentID_FaceVerification\student-id-face-matching\api\List of candidates\extracted_list\student_list_from_excel.txt', 'w', encoding='utf-8') as f:
        for student in students_from_excel:
            f.write(student + '\n')
    print("Danh sách sinh viên từ Excel đã được lưu vào file 'student_list_from_excel.txt'.")
