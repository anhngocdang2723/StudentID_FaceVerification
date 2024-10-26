from paddleocr import PaddleOCR
import os
import openpyxl
from unidecode import unidecode

ocr = PaddleOCR(use_angle_cls=True, lang='vi')

excel_path = r'D:\Edu\Python\StudentID_FaceVerification\student-id-face-matching\api\List of candidates\DemoDanhSach.xlsx'

def find_columns(sheet):
    name_col = None
    msv_col = None

    for row in sheet.iter_rows(values_only=True):
        for idx, cell_value in enumerate(row):
            if cell_value:
                if "họ tên" in str(cell_value).lower():
                    name_col = idx + 1
                elif "mã sinh viên" in str(cell_value).lower():
                    msv_col = idx + 1                
                if name_col and msv_col:
                    return name_col, msv_col
    return name_col, msv_col

def read_from_excel(excel_path):
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        name_col, msv_col = find_columns(sheet)

        if name_col and msv_col:
            excel_data = []
            for row in sheet.iter_rows(min_row=sheet.min_row + 1, values_only=True):
                name = row[name_col - 1]
                msv = row[msv_col - 1]
                if name and msv: 
                    name = unidecode(name).upper()
                    excel_data.append(f"{name} - {msv}")
            return excel_data
        else:
            print("Không tìm thấy cột 'Họ tên' và 'Mã sinh viên'.")
            return []
    else:
        print(f"File Excel {excel_path} không tồn tại.")
        return []

students_from_excel = read_from_excel(excel_path)

######## Sẽ cập nhật lưu kết quả đọc được từ file excel vào DB sau ########
# hiển thị và lưu kết quả
if students_from_excel:
    print("Danh sách sinh viên từ Excel:")
    for student in students_from_excel:
        print(student)

    with open(r'D:\Edu\Python\StudentID_FaceVerification\student-id-face-matching\api\List of candidates\extracted_list\student_list_from_excel.txt', 'w', encoding='utf-8') as f:
        for student in students_from_excel:
            f.write(student + '\n')
    print("Danh sách sinh viên từ Excel đã được lưu vào file 'student_list_from_excel.txt'.")
