# excel_reader.py
import os
import openpyxl
from unidecode import unidecode
from paddleocr import PaddleOCR

ocr = PaddleOCR(use_angle_cls=True, lang='vi')

def find_columns(sheet):
    name_col = None
    msv_col = None

    for row in sheet.iter_rows(values_only=True):
        for idx, cell_value in enumerate(row):
            if cell_value:
                if "họ tên" in str(cell_value).lower():
                    name_col = idx + 1
                elif "mã sinh viên" in str(cell_value).lower():
                    msv_col = idx + 1                
                if name_col and msv_col:
                    return name_col, msv_col
    return name_col, msv_col

def read_from_excel(excel_path):
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        name_col, msv_col = find_columns(sheet)

        if name_col and msv_col:
            excel_data = []
            for row in sheet.iter_rows(min_row=sheet.min_row + 1, values_only=True):
                name = row[name_col - 1]
                msv = row[msv_col - 1]
                if name and msv: 
                    name = unidecode(name).upper()
                    excel_data.append(f"{name} - {msv}")
            return excel_data
        else:
            print("Không tìm thấy cột 'Họ tên' và 'Mã sinh viên'.")
            return []
    else:
        print(f"File Excel {excel_path} không tồn tại.")
        return []

def save_students_to_file(students, output_path):
    with open(output_path, 'w', encoding='utf-8') as f:
        for student in students:
            f.write(student + '\n')
    print(f"Danh sách sinh viên từ Excel đã được lưu vào file '{output_path}'.")


# if __name__ == "__main__":
#     excel_file_path = r"D:\Edu\Python\StudentID_FaceVerification\student-id-face-matching\api\List of candidates\students.xlsx"  # Đường dẫn đến file Excel mẫu
#     output_file_path = r"D:\Edu\Python\StudentID_FaceVerification\student-id-face-matching\api\List of candidates\extracted_list\students.txt"  # Đường dẫn đến file văn bản đầu ra

#     # Đọc dữ liệu từ file Excel
#     students = read_from_excel(excel_file_path)

#     # Nếu có dữ liệu, lưu vào file
#     if students:
#         save_students_to_file(students, output_file_path)
